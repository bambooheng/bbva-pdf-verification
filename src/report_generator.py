"""
报告生成模块：生成 JSON、Markdown 和 Excel 格式的审计报告
"""
import json
import logging
import pandas as pd
from pathlib import Path
from typing import Dict, Any, List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ReportGenerator:
    """报告生成器"""
    
    def __init__(self, output_json_path: str, output_markdown_path: str, output_excel_path: str = None):
        """
        初始化报告生成器
        
        Args:
            output_json_path: JSON 报告输出路径
            output_markdown_path: Markdown 报告输出路径
            output_excel_path: Excel 报告输出路径（可选）
        """
        self.output_json_path = Path(output_json_path)
        self.output_markdown_path = Path(output_markdown_path)
        self.output_excel_path = Path(output_excel_path) if output_excel_path else None
        
        # 确保输出目录存在
        self.output_json_path.parent.mkdir(parents=True, exist_ok=True)
        self.output_markdown_path.parent.mkdir(parents=True, exist_ok=True)
        if self.output_excel_path:
            self.output_excel_path.parent.mkdir(parents=True, exist_ok=True)
    
    def generate_json_report(self, results: List[Dict[str, Any]], summary: Dict[str, Any], 
                            metadata: Dict[str, Any] = None) -> str:
        """
        生成 JSON 格式报告
        
        Args:
            results: 审计结果列表
            summary: 摘要统计
            metadata: 元数据（可选）
            
        Returns:
            生成的 JSON 文件路径
        """
        report = {
            "metadata": {
                "generated_at": datetime.now().isoformat(),
                "total_rules": summary.get("total_rules", 0),
                **(metadata or {})
            },
            "summary": summary,
            "results": results
        }
        
        try:
            with open(self.output_json_path, 'w', encoding='utf-8') as f:
                json.dump(report, f, ensure_ascii=False, indent=2)
            
            logger.info(f"JSON 报告已生成: {self.output_json_path}")
            return str(self.output_json_path)
        
        except Exception as e:
            logger.error(f"生成 JSON 报告失败: {e}")
            raise
    
    def generate_markdown_report(self, results: List[Dict[str, Any]], summary: Dict[str, Any],
                                 metadata: Dict[str, Any] = None) -> str:
        """
        生成 Markdown 格式报告
        
        Args:
            results: 审计结果列表
            summary: 摘要统计
            metadata: 元数据（可选）
            
        Returns:
            生成的 Markdown 文件路径
        """
        lines = []
        
        # 标题
        lines.append("# BBVA 银行流水审计报告\n")
        lines.append(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        
        # 元数据
        if metadata:
            lines.append("## 元数据\n")
            for key, value in metadata.items():
                lines.append(f"- **{key}**: {value}")
            lines.append("")
        
        # 摘要统计
        lines.append("## 审计摘要\n")
        lines.append(f"- **总规则数**: {summary.get('total_rules', 0)}")
        lines.append(f"- **命中规则数** (违规): {summary.get('hit_count', 0)}")
        lines.append(f"- **未命中规则数** (合规): {summary.get('not_hit_count', 0)}")
        lines.append(f"- **无法判断规则数**: {summary.get('unknown_count', 0)}")
        lines.append("")
        lines.append("### 置信度分布\n")
        lines.append(f"- **高置信度**: {summary.get('high_confidence_count', 0)}")
        lines.append(f"- **中置信度**: {summary.get('medium_confidence_count', 0)}")
        lines.append(f"- **低置信度**: {summary.get('low_confidence_count', 0)}")
        lines.append("")
        
        # 详细结果
        lines.append("## 详细审计结果\n")
        
        for idx, result in enumerate(results, 1):
            lines.append(f"### {idx}. {result['rule_name']}\n")
            lines.append(f"**规则ID**: `{result['rule_id']}`\n")
            
            # 命中状态
            hit_status = result.get('hit')
            if hit_status is True:
                hit_display = "✅ **命中** (违规)"
            elif hit_status is False:
                hit_display = "✅ **未命中** (合规)"
            else:
                hit_display = "❓ **无法判断**"
            
            lines.append(f"**命中状态**: {hit_display}\n")
            lines.append(f"**置信度**: {result.get('confidence', 'unknown').upper()}\n")
            
            # 判断依据
            evidence = result.get('evidence', '')
            if evidence:
                lines.append(f"**判断依据**:\n")
                lines.append(f"> {evidence}\n")
            
            # 补充说明
            notes = result.get('notes', '')
            if notes:
                lines.append(f"**补充说明**: {notes}\n")
            
            lines.append("---\n")
        
        # 写入文件
        try:
            with open(self.output_markdown_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(lines))
            
            logger.info(f"Markdown 报告已生成: {self.output_markdown_path}")
            return str(self.output_markdown_path)
        
        except Exception as e:
            logger.error(f"生成 Markdown 报告失败: {e}")
            raise
    
    def generate_excel_report(self, results: List[Dict[str, Any]], summary: Dict[str, Any],
                             metadata: Dict[str, Any] = None) -> str:
        """
        生成 Excel 格式报告
        
        Args:
            results: 审计结果列表
            summary: 摘要统计
            metadata: 元数据（可选）
            
        Returns:
            生成的 Excel 文件路径
        """
        if not self.output_excel_path:
            raise ValueError("未配置 Excel 输出路径")
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "审计报告"
            
            # 样式定义
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            title_font = Font(bold=True, size=14)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            row = 1
            
            # 标题
            ws.merge_cells(f'A{row}:F{row}')
            title_cell = ws[f'A{row}']
            title_cell.value = "BBVA 银行流水审计报告"
            title_cell.font = title_font
            title_cell.alignment = center_alignment
            row += 1
            
            # 生成时间
            ws.merge_cells(f'A{row}:F{row}')
            time_cell = ws[f'A{row}']
            time_cell.value = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            time_cell.alignment = center_alignment
            row += 2
            
            # 元数据
            if metadata:
                ws.merge_cells(f'A{row}:F{row}')
                ws[f'A{row}'].value = "元数据"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                for key, value in metadata.items():
                    ws[f'A{row}'].value = f"{key}:"
                    ws[f'B{row}'].value = value
                    row += 1
                row += 1
            
            # 摘要统计
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'].value = "审计摘要"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            summary_data = [
                ["总规则数", summary.get('total_rules', 0)],
                ["命中规则数 (违规)", summary.get('hit_count', 0)],
                ["未命中规则数 (合规)", summary.get('not_hit_count', 0)],
                ["无法判断规则数", summary.get('unknown_count', 0)],
                ["高置信度", summary.get('high_confidence_count', 0)],
                ["中置信度", summary.get('medium_confidence_count', 0)],
                ["低置信度", summary.get('low_confidence_count', 0)],
            ]
            
            for item in summary_data:
                ws[f'A{row}'].value = item[0]
                ws[f'B{row}'].value = item[1]
                row += 1
            
            row += 1
            
            # 详细结果表头
            headers = ["规则ID", "规则名称", "命中状态", "置信度", "判断依据", "补充说明"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border
            
            row += 1
            
            # 详细结果数据
            for result in results:
                # 规则ID
                ws.cell(row=row, column=1).value = result.get('rule_id', '')
                ws.cell(row=row, column=1).border = border
                ws.cell(row=row, column=1).alignment = left_alignment
                
                # 规则名称
                ws.cell(row=row, column=2).value = result.get('rule_name', '')
                ws.cell(row=row, column=2).border = border
                ws.cell(row=row, column=2).alignment = left_alignment
                
                # 命中状态
                hit_status = result.get('hit')
                if hit_status is True:
                    hit_display = "命中 (违规)"
                    hit_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                elif hit_status is False:
                    hit_display = "未命中 (合规)"
                    hit_fill = PatternFill(start_color="51CF66", end_color="51CF66", fill_type="solid")
                else:
                    hit_display = "无法判断"
                    hit_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                
                hit_cell = ws.cell(row=row, column=3)
                hit_cell.value = hit_display
                hit_cell.fill = hit_fill
                hit_cell.border = border
                hit_cell.alignment = center_alignment
                hit_cell.font = Font(bold=True)
                
                # 置信度
                confidence = result.get('confidence', 'unknown').upper()
                ws.cell(row=row, column=4).value = confidence
                ws.cell(row=row, column=4).border = border
                ws.cell(row=row, column=4).alignment = center_alignment
                
                # 判断依据
                evidence = result.get('evidence', '')
                ws.cell(row=row, column=5).value = evidence
                ws.cell(row=row, column=5).border = border
                ws.cell(row=row, column=5).alignment = left_alignment
                
                # 补充说明
                notes = result.get('notes', '')
                ws.cell(row=row, column=6).value = notes
                ws.cell(row=row, column=6).border = border
                ws.cell(row=row, column=6).alignment = left_alignment
                
                row += 1
            
            # 调整列宽
            ws.column_dimensions['A'].width = 25  # 规则ID
            ws.column_dimensions['B'].width = 25  # 规则名称
            ws.column_dimensions['C'].width = 18  # 命中状态
            ws.column_dimensions['D'].width = 12  # 置信度
            ws.column_dimensions['E'].width = 60  # 判断依据
            ws.column_dimensions['F'].width = 40  # 补充说明
            
            # 设置行高
            for r in range(1, row):
                ws.row_dimensions[r].height = 20
            
            # 保存文件
            wb.save(self.output_excel_path)
            
            logger.info(f"Excel 报告已生成: {self.output_excel_path}")
            return str(self.output_excel_path)
        
        except Exception as e:
            logger.error(f"生成 Excel 报告失败: {e}")
            raise
    
    def generate_all_reports(self, results: List[Dict[str, Any]], summary: Dict[str, Any],
                             metadata: Dict[str, Any] = None) -> tuple:
        """
        同时生成 JSON、Markdown 和 Excel 报告
        
        Args:
            results: 审计结果列表
            summary: 摘要统计
            metadata: 元数据（可选）
            
        Returns:
            (json_path, markdown_path, excel_path) 元组
        """
        json_path = self.generate_json_report(results, summary, metadata)
        markdown_path = self.generate_markdown_report(results, summary, metadata)
        
        excel_path = None
        if self.output_excel_path:
            excel_path = self.generate_excel_report(results, summary, metadata)
        
        return json_path, markdown_path, excel_path
    
    def generate_both_reports(self, results: List[Dict[str, Any]], summary: Dict[str, Any],
                             metadata: Dict[str, Any] = None) -> tuple:
        """
        同时生成 JSON 和 Markdown 报告（保持向后兼容）
        
        Args:
            results: 审计结果列表
            summary: 摘要统计
            metadata: 元数据（可选）
            
        Returns:
            (json_path, markdown_path) 元组
        """
        json_path = self.generate_json_report(results, summary, metadata)
        markdown_path = self.generate_markdown_report(results, summary, metadata)
        
        return json_path, markdown_path

